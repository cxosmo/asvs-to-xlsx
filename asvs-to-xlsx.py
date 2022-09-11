#!/usr/bin/env python3

import argparse
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.workbook import Workbook
import string


def json_to_dict(filename):
    """
    Read JSON file and return as dictionary
    """
    try:
        with open(filename) as f:
            data = json.load(f)
            return data
    except Exception as e:
        print(f"Unable to open {filename}: {e}")


def format_sheet(workbook):
    """Iterates over sheets in workbook, formatting cells based on row position."""
    uppercase_alphabet = string.ascii_uppercase
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(name="Calibri", size=11, bold=True)
    blue_fill = PatternFill(fill_type="solid", fgColor="0099CCFF")
    grey_fill = PatternFill(fill_type="solid", fgColor="00F1EDED")
    alignment = Alignment(horizontal="left")
    regular_font = Font(name="Calibri", size=10)
    for sheet in workbook:
        for row in sheet.rows:
            for cell in row:
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = blue_fill
                    cell.border = thin_border
                    cell.alignment = alignment
                else:
                    cell.font = regular_font
                    cell.border = thin_border
                    cell.alignment = alignment
                    cell.fill = grey_fill
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 0.8
            sheet.column_dimensions[column].width = adjusted_width


def create_workbook(json, custom_output_name, custom_columns):
    """Creates workbook based upon ASVS JSON input + (optional) custom naming and column values."""
    wb = Workbook()
    del wb["Sheet"]
    workbook_title = f"{json['ShortName']}-{json['Version']}.xlsx"
    if custom_output_name:
        workbook_title = custom_output_name
    for i in range(len(json["Requirements"])):
        sheet_name = json["Requirements"][i]["Name"]
        wb.create_sheet(sheet_name)
        wb.active = wb[sheet_name]
        ws = wb.active
        header_list = ["Category", "#", "CWE", "Description"] + custom_columns
        ws.append(header_list)
        for category in json["Requirements"][i]["Items"]:
            for item in category["Items"]:
                try:
                    ws.append(
                        [
                            category["Name"],
                            item["Shortcode"],
                            item["CWE"][0],
                            item["Description"],
                        ]
                    )
                except Exception as e:
                    ws.append(
                        [category["Name"], item["Shortcode"], None, item["Description"]]
                    )
    format_sheet(wb)
    wb.save(filename=workbook_title)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-i",
        "--input_file",
        help="Name of ASVS JSON file to parse (e.g. 'OWASP Application Security Verification Standard 4.0.3-en.json').",
        required="True",
        action="store",
    )
    parser.add_argument(
        "-o",
        "--output_file",
        help="Filename for xlsx output (optional; defaults to ASVS-n.xlsx).",
        default=False,
        action="store",
    )
    parser.add_argument(
        "-c",
        "--columns",
        help="Additional custom column(s) to include in xlsx output, supporting multiple uses (e.g. -c 'Comments' -c 'Findings').",
        nargs="+",
        action="append",
        default=[],
    )

    args = parser.parse_args()
    args.columns = [val for sublist in args.columns for val in sublist]
    create_workbook(json_to_dict(args.input_file), args.output_file, args.columns)
